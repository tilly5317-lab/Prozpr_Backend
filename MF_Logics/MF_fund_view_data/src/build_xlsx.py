"""Orchestrator: run Tier 1+2+3 for a list of schemes and assemble the xlsx
matching the MF_evaluation framework column order.
"""
from __future__ import annotations

import argparse
import json
import math
import time
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .tier1_derive import derive_row
from .tier2_returns import compute as tier2_compute
from .tier3_groww import extract_tier3, fetch_for_scheme

ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "build" / "output"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
LOGS_DIR = ROOT / "build" / "logs"
LOGS_DIR.mkdir(parents=True, exist_ok=True)

# Final column order (matches the framework xlsx)
SECTION_GROUPS = [
    ("1. Fund overview", [
        "schemeCode", "schemeName", "isin_growth", "isin_div_reinvest",
        "asset_class", "asset_subcategory",
        "plan_class", "div_or_growth", "investor", "min_investment", "asset_size_cr",
    ]),
    ("2. Quantitative — Fund Returns", [
        "1Y CAGR", "3Y CAGR", "5Y CAGR", "7Y CAGR",
        "3Y rolling 2020", "3Y rolling 2021", "3Y rolling 2022", "3Y rolling 2023",
        "3Y rolling 2024", "3Y rolling 2025", "3Y rolling 2026",
        "%ile rank 3Y rolling 2020", "%ile rank 3Y rolling 2021", "%ile rank 3Y rolling 2022",
        "%ile rank 3Y rolling 2023", "%ile rank 3Y rolling 2024", "%ile rank 3Y rolling 2025",
        "%ile rank 3Y rolling 2026",
    ]),
    ("Benchmark returns", [
        "Bench 3Y rolling 2020", "Bench 3Y rolling 2021", "Bench 3Y rolling 2022",
        "Bench 3Y rolling 2023", "Bench 3Y rolling 2024", "Bench 3Y rolling 2025",
        "Bench 3Y rolling 2026",
        "Sharpe (3Y)", "Beta", "Tracking Error",
    ]),
    ("3. Portfolio Quality", [
        "size_exposure_LMS", "style_exposure_GV", "top10_holdings_weight_pct",
        "alpha_selection_l3y_pct", "alpha_allocation_l3y_pct", "portfolio_churn_l3y_pct",
    ]),
    ("4. Costs & Efficiency", [
        "expense_ratio_pct", "performance_fees", "entry_load", "exit_load", "exit_load_period",
    ]),
    ("5. Team", [
        "lead_pm_name", "manager_start_date", "pm_turnover_l5y", "parent_fund_house",
    ]),
    ("6. Tax", [
        "st_rate", "st_period", "lt_rate", "lt_period",
    ]),
    ("Diagnostics", [
        "benchmark_token", "groww_benchmark", "groww_isin_match", "_groww_slug",
        "nav_start", "nav_end", "nav_points",
    ]),
]


def _ordered_columns() -> list[str]:
    cols = []
    for _, fields in SECTION_GROUPS:
        cols.extend(fields)
    return cols


def process_one(row: dict, log) -> dict:
    code = int(row["schemeCode"])
    name = row["schemeName"]
    sub = row.get("sub_category", "")
    out = {"schemeCode": code, "schemeName": name}
    out.update(derive_row(row))
    try:
        t2 = tier2_compute(code, sub, name)
        out.update(t2)
    except Exception as e:
        log.write(f"{code} TIER2_ERR {e}\n")
    try:
        detail = fetch_for_scheme(code, name)
        t3 = extract_tier3(detail, (out.get("isin_growth", ""), out.get("isin_div_reinvest", "")))
        out.update(t3)
    except Exception as e:
        log.write(f"{code} TIER3_ERR {e}\n")
    out["alpha_selection_l3y_pct"] = None
    out["alpha_allocation_l3y_pct"] = None
    return out


def add_percentile_ranks(df: pd.DataFrame) -> pd.DataFrame:
    """Compute %ile rank within (asset_subcategory, year) for each rolling year.

    %ile = percent of peers (in same sub_category) with strictly lower 3Y rolling return.
    Higher = better.
    """
    years = [2020, 2021, 2022, 2023, 2024, 2025, 2026]
    for y in years:
        col = f"3Y rolling {y}"
        rank_col = f"%ile rank 3Y rolling {y}"
        df[rank_col] = None
        for sub, grp in df.groupby("asset_subcategory"):
            vals = grp[col].dropna()
            if len(vals) < 3:
                continue
            ranks = vals.rank(method="average", pct=True) * 100
            df.loc[ranks.index, rank_col] = ranks.round(1)
    return df


def build_xlsx(funds: list[dict], output_path: Path, source_csv: Path | None = None) -> None:
    df = pd.DataFrame(funds)
    df = add_percentile_ranks(df)

    cols = ["schemeCode", "schemeName"] + _ordered_columns()
    cols = [c for c in cols if c not in {"schemeCode", "schemeName"} or cols.index(c) <= 1]
    # Reorder: final list is exactly _ordered_columns() (which already includes scheme cols)
    cols = _ordered_columns()
    for c in cols:
        if c not in df.columns:
            df[c] = None
    df = df[cols]

    wb = Workbook()
    ws = wb.active
    ws.title = "All funds"

    # Header rows: section labels (row 1), field names (row 2)
    section_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    field_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    bold = Font(bold=True)
    centered = Alignment(horizontal="center", vertical="center", wrap_text=True)

    col_idx = 1
    for section, fields in SECTION_GROUPS:
        start = col_idx
        for f in fields:
            ws.cell(row=2, column=col_idx, value=f).fill = field_fill
            ws.cell(row=2, column=col_idx).font = bold
            ws.cell(row=2, column=col_idx).alignment = centered
            col_idx += 1
        end = col_idx - 1
        if end >= start:
            ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
            c = ws.cell(row=1, column=start, value=section)
            c.font = Font(bold=True)
            c.alignment = centered
            c.fill = section_fill

    # Data rows
    for r_idx, (_, row) in enumerate(df.iterrows(), start=3):
        for c_idx, col in enumerate(cols, start=1):
            v = row[col]
            if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
                v = None
            if isinstance(v, float):
                ws.cell(row=r_idx, column=c_idx, value=round(v, 4))
            else:
                ws.cell(row=r_idx, column=c_idx, value=v)

    # Column widths
    for c_idx, col in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = max(12, min(28, len(col) + 4))

    ws.freeze_panes = "C3"

    # Coverage sheet
    cov = wb.create_sheet("Coverage")
    cov.append(["field", "filled", "total", "pct"])
    for c in cols:
        filled = df[c].notna().sum()
        total = len(df)
        cov.append([c, int(filled), int(total), round(100 * filled / total, 1) if total else 0])
    for col_letter in ("A", "B", "C", "D"):
        cov.column_dimensions[col_letter].width = 30 if col_letter == "A" else 12

    # Benchmark map sheet (informational)
    from .benchmarks import SUBCAT_BENCHMARK
    bm = wb.create_sheet("Benchmark map")
    bm.append(["sub_category", "benchmark_token"])
    for k, v in SUBCAT_BENCHMARK.items():
        bm.append([k, str(v)])
    bm.column_dimensions["A"].width = 40
    bm.column_dimensions["B"].width = 18

    wb.save(output_path)
    print(f"Wrote {output_path} ({len(df)} rows)")


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--csv", default="latest_nav_active.csv")
    p.add_argument("--codes", nargs="+", type=int, help="Subset of scheme_codes (pilot)")
    p.add_argument("--out", default="build/output/mf_evaluation.xlsx")
    p.add_argument("--checkpoint", default="build/output/_partial.csv",
                   help="Incremental CSV checkpoint (resumable)")
    p.add_argument("--checkpoint-every", type=int, default=100)
    args = p.parse_args()

    df = pd.read_csv(args.csv)
    if args.codes:
        df = df[df["schemeCode"].isin(args.codes)].copy()
    df = df.reset_index(drop=True)
    print(f"Processing {len(df)} schemes…")

    # Resume from checkpoint if present
    ckpt = ROOT / args.checkpoint
    funds = []
    done_codes: set[int] = set()
    if ckpt.exists():
        prev = pd.read_csv(ckpt)
        funds = prev.to_dict("records")
        done_codes = {int(c) for c in prev["schemeCode"].tolist()}
        print(f"  resume: {len(done_codes)} already in checkpoint")

    log_path = LOGS_DIR / f"run_{int(time.time())}.log"
    with open(log_path, "w") as log:
        for i, (_, row) in enumerate(df.iterrows()):
            code = int(row["schemeCode"])
            if code in done_codes:
                continue
            t0 = time.time()
            try:
                res = process_one(row.to_dict(), log)
            except Exception as e:
                log.write(f"{code} FATAL {e}\n")
                res = {"schemeCode": code, "schemeName": row["schemeName"], "_error": str(e)}
            funds.append(res)
            elapsed = time.time() - t0
            if i % 25 == 0 or elapsed > 3:
                print(f"  [{i+1}/{len(df)}] {code} {row['schemeName'][:55]:55s}  {elapsed:.1f}s")
            if (len(funds) % args.checkpoint_every) == 0:
                pd.DataFrame(funds).to_csv(ckpt, index=False)

    pd.DataFrame(funds).to_csv(ckpt, index=False)
    out_path = ROOT / args.out
    build_xlsx(funds, out_path)
    print(f"Log: {log_path}")


if __name__ == "__main__":
    main()
