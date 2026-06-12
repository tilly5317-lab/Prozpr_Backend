"""Issue-report storage + notification: Google Sheet register, screenshots, email.

There is deliberately NO database table for issue reports. The shared Google
Sheet IS the issue register, with a local xlsx as automatic fallback:

- ``ISSUE_SHEET_WEBHOOK_URL`` (.env) — a Google Apps Script web app attached to
  the admin's Google Sheet; we POST one row per report:
  Date | User Name | Email | Source | Issue | Screenshot | Remarks
  (Remarks stays empty — it is the admin's column to fill while triaging.)
  ``ISSUE_SHEET_TOKEN`` is a shared secret the script checks so strangers
  cannot post junk rows if the URL leaks.
- ``ISSUE_REPORTS_XLSX`` (default ``Prozpr_Backend/issue_reports/issue_reports.xlsx``)
  — local fallback register, used when the webhook is unset or unreachable so
  a report is never lost.
- ``issue_reports/screenshots/<report_id>.<ext>`` — optional screenshot file
  (also attached to the email).
- Email to ``SUPPORT_EMAIL_TO`` via Zoho SMTP (``smtp.zoho.com:465`` SSL).
  Skipped with a log warning when ``SMTP_PASSWORD`` is not configured — a
  report must never fail because mail is down.

All functions are synchronous on purpose: the router runs the register append /
screenshot save via ``asyncio.to_thread`` and the email as a background task.
"""

from __future__ import annotations

import logging
import smtplib
import ssl
import threading
from datetime import datetime, timedelta, timezone
from email.message import EmailMessage
from pathlib import Path

import httpx
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.core.config import get_settings

logger = logging.getLogger(__name__)

try:
    from zoneinfo import ZoneInfo

    IST = ZoneInfo("Asia/Kolkata")
except Exception:  # Windows without the `tzdata` package — fixed offset is exact for IST.
    IST = timezone(timedelta(hours=5, minutes=30), "IST")

# Prozpr_Backend/issue_reports/ — alongside app/, not inside it.
_REPORTS_DIR = Path(__file__).resolve().parents[4] / "issue_reports"
SCREENSHOTS_DIR = _REPORTS_DIR / "screenshots"
DEFAULT_EXCEL_PATH = _REPORTS_DIR / "issue_reports.xlsx"

EXCEL_HEADERS = ["Date", "User Name", "Email", "Source", "Issue", "Screenshot", "Remarks"]

ALLOWED_SOURCES = {
    "Chat Response",
    "Portfolio NAV",
    "Rebalancing",
    "Goal Planning",
    "Onboarding",
    "Other",
}

ALLOWED_IMAGE_TYPES = {
    "image/png": ".png",
    "image/jpeg": ".jpg",
    "image/webp": ".webp",
    "image/gif": ".gif",
}
MAX_SCREENSHOT_BYTES = 5 * 1024 * 1024  # 5 MB

# openpyxl read-modify-write is not safe under concurrent appends.
_excel_lock = threading.Lock()


def get_excel_path() -> Path:
    """Fallback workbook location; override with ``ISSUE_REPORTS_XLSX`` in .env."""
    raw = (get_settings().get_issue_reports_xlsx() or "").strip()
    return Path(raw) if raw else DEFAULT_EXCEL_PATH


def append_to_google_sheet(
    created_at: datetime,
    user_name: str,
    user_email: str,
    source_label: str,
    description: str,
    screenshot_name: str | None,
) -> bool:
    """POST one register row to the Apps Script webhook.

    Returns False when no webhook is configured; raises on HTTP/script errors
    so the router can fall back to the local xlsx register.
    """
    settings = get_settings()
    url = settings.get_issue_sheet_webhook_url()
    if not url:
        return False

    payload = {
        "token": settings.get_issue_sheet_token() or "",
        "date": created_at.astimezone(IST).strftime("%Y-%m-%d %H:%M:%S"),
        "user_name": user_name,
        "email": user_email,
        "source": source_label,
        "issue": description,
        "screenshot": screenshot_name or "",
    }
    # Apps Script answers with a 302 to script.googleusercontent.com — follow it.
    resp = httpx.post(url, json=payload, timeout=20, follow_redirects=True)
    resp.raise_for_status()
    try:
        body = resp.json()
    except ValueError as exc:
        raise RuntimeError(f"Issue sheet webhook returned non-JSON: {resp.text[:200]}") from exc
    if body.get("ok") is not True:
        raise RuntimeError(f"Issue sheet webhook rejected the row: {body}")
    return True


def save_screenshot(content: bytes, content_type: str, report_id: str) -> str:
    """Persist screenshot bytes; returns the stored path as a string."""
    ext = ALLOWED_IMAGE_TYPES[content_type]
    SCREENSHOTS_DIR.mkdir(parents=True, exist_ok=True)
    path = SCREENSHOTS_DIR / f"{report_id}{ext}"
    path.write_bytes(content)
    return str(path)


def _new_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Issue Reports"
    ws.append(EXCEL_HEADERS)
    for col_idx in range(1, len(EXCEL_HEADERS) + 1):
        ws.cell(row=1, column=col_idx).font = Font(bold=True)
    widths = [20, 24, 30, 22, 60, 28, 40]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    return wb


def append_to_excel(
    created_at: datetime,
    user_name: str,
    user_email: str,
    source_label: str,
    description: str,
    screenshot_name: str | None,
) -> None:
    """Append one report row to the issue register (first column = date).

    Raises on I/O errors (e.g. the workbook is open in Excel and locked) so
    the router can surface a clear message instead of silently dropping data.
    """
    excel_path = get_excel_path()
    with _excel_lock:
        excel_path.parent.mkdir(parents=True, exist_ok=True)
        if excel_path.exists():
            wb = load_workbook(excel_path)
        else:
            wb = _new_workbook()
        ws = wb.active
        ws.append([
            created_at.astimezone(IST).strftime("%Y-%m-%d %H:%M:%S"),
            user_name,
            user_email,
            source_label,
            description,
            screenshot_name or "",
            "",  # Remarks — filled in by the admin while triaging
        ])
        wb.save(excel_path)


def send_issue_email(
    created_at: datetime,
    user_name: str,
    user_email: str,
    source_label: str,
    description: str,
    screenshot: bytes | None,
    screenshot_type: str | None,
) -> None:
    """Email the report to the support inbox; failures are logged, never raised."""
    settings = get_settings()
    password = settings.get_smtp_password()
    if not password:
        logger.warning(
            "SMTP_PASSWORD not set — issue report email skipped (report is in the Excel log)."
        )
        return

    sender = settings.get_smtp_user()
    recipient = settings.get_support_email_to()
    when = created_at.astimezone(IST).strftime("%d %b %Y, %I:%M %p IST")

    msg = EmailMessage()
    msg["Subject"] = f"[Prozpr Issue] {source_label} — {user_name or user_email or 'Unknown user'}"
    msg["From"] = sender
    msg["To"] = recipient
    msg.set_content(
        f"A user reported an issue on Prozpr.\n"
        f"\n"
        f"Date       : {when}\n"
        f"User Name  : {user_name or '-'}\n"
        f"Email      : {user_email or '-'}\n"
        f"Source     : {source_label}\n"
        f"\n"
        f"Issue:\n{description}\n"
        f"\n"
        f"Screenshot : {'attached' if screenshot else 'not provided'}\n"
    )

    if screenshot and screenshot_type:
        maintype, _, subtype = screenshot_type.partition("/")
        ext = ALLOWED_IMAGE_TYPES.get(screenshot_type, ".png")
        msg.add_attachment(
            screenshot, maintype=maintype, subtype=subtype, filename=f"screenshot{ext}"
        )

    host = settings.get_smtp_host()
    port = settings.get_smtp_port()
    try:
        if port == 465:
            with smtplib.SMTP_SSL(host, port, context=ssl.create_default_context(), timeout=30) as smtp:
                smtp.login(sender, password)
                smtp.send_message(msg)
        else:
            with smtplib.SMTP(host, port, timeout=30) as smtp:
                smtp.starttls(context=ssl.create_default_context())
                smtp.login(sender, password)
                smtp.send_message(msg)
        logger.info("Issue report email sent to %s (source=%s)", recipient, source_label)
    except Exception:
        logger.exception("Failed to send issue report email (report is in the Excel log).")
