"""DEV-ONLY: Extract baseline scenario from Sourabh's Excel for cashflow_statement parity.

Reads `Local_logics/Sourabh_Logics/goal_based_allocation_model (10).xlsx`, sheet
`Goal planning`, and produces:

  AI_Agents/src/cashflow_statement/tests/fixtures/excel_reference/baseline/
    input.json     # loadable as GoalPlanningInput
    expected.json  # flat dict: cell-or-path → expected value

Cell mapping is documented in
  AI_Agents/src/cashflow_statement/tests/fixtures/excel_reference/cell_mapping.md

Run from repo root:
    .venv-mac/bin/python scripts/extract_excel_reference.py
"""
from __future__ import annotations

import json
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


REPO_ROOT = Path(__file__).resolve().parent.parent
EXCEL_PATH = (
    REPO_ROOT.parent.parent
    / "Local_logics" / "Sourabh_Logics"
    / "goal_based_allocation_model (10).xlsx"
)
OUT_DIR = (
    REPO_ROOT
    / "AI_Agents/src/cashflow_statement/tests/fixtures/excel_reference/baseline"
)


# ---------------------------------------------------------------------------
# Cell helpers
# ---------------------------------------------------------------------------

def _v(ws: Worksheet, cell: str):
    return ws[cell].value


def _date(ws: Worksheet, cell: str) -> date | None:
    v = ws[cell].value
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    raise TypeError(f"{cell}: expected date, got {type(v).__name__}: {v!r}")


def _float(ws: Worksheet, cell: str) -> float | None:
    v = ws[cell].value
    if v is None or v == "":
        return None
    return float(v)


def _str(ws: Worksheet, cell: str) -> str | None:
    v = ws[cell].value
    if v is None:
        return None
    return str(v).strip()


def _bool(ws: Worksheet, cell: str) -> bool | None:
    v = ws[cell].value
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, str):
        return v.strip().lower() in ("true", "yes", "1")
    return bool(v)


# ---------------------------------------------------------------------------
# Input extraction
# ---------------------------------------------------------------------------

def extract_assumptions(ws: Worksheet) -> dict:
    return {
        "inflation_property": _float(ws, "B3"),
        "inflation_child_abroad_education": _float(ws, "B4"),
        "inflation_child_local_education": _float(ws, "B5"),
        "inflation_child_marriage": _float(ws, "B6"),
        "inflation_household_expense": _float(ws, "B10"),
        "annual_income_growth": _float(ws, "B8"),
        "annual_invested_amount_growth": _float(ws, "B9"),
        "roi_near_term_post_tax": _float(ws, "B11"),
        "roi_mid_term_post_tax": _float(ws, "B12"),
        "roi_long_term_post_tax": _float(ws, "B13"),
        "roi_retired_portfolio_annual": _float(ws, "B14"),
    }


def extract_profile(ws: Worksheet) -> dict:
    return {
        "latest_update_date": _date(ws, "B18").isoformat(),
        "annual_income": _float(ws, "B22"),
        "tax_rate": _float(ws, "B23"),
        "financial_assets": _float(ws, "B24"),
        "financial_liabilities_excl_mortgage": _float(ws, "B25"),
        "monthly_household_expense": _float(ws, "B40"),
        "monthly_investment_next_12m": _float(ws, "B27"),
    }


def extract_retirement(ws: Worksheet) -> dict:
    out = {
        "date_of_birth": _date(ws, "B34").isoformat(),
        "retirement_age": int(_float(ws, "B35")),
        "assumed_total_age": int(_float(ws, "B36")),
    }
    user_date = _date(ws, "B38")
    if user_date is not None:
        out["retirement_date_override"] = user_date.isoformat()
    user_pv = _float(ws, "B44")
    if user_pv is not None:
        out["retirement_corpus_pv_override"] = user_pv
    return out


def extract_current_properties(ws: Worksheet) -> list[dict]:
    """Cols B-F (2-6), rows 49-60. Reads annualized rate from row 58 (= Excel's
    B58:F58, computed via simple annualization from monthly rate B57)."""
    out: list[dict] = []
    for col in range(2, 7):
        owns = ws.cell(row=50, column=col).value
        if owns is None or owns is False or owns == "":
            continue
        name = ws.cell(row=49, column=col).value
        if name is None:
            continue
        has_mortgage = bool(ws.cell(row=51, column=col).value)
        balance = ws.cell(row=52, column=col).value
        emi = ws.cell(row=56, column=col).value
        annual_rate = ws.cell(row=58, column=col).value
        prop = {
            "name": str(name),
            "has_mortgage": has_mortgage,
        }
        if has_mortgage:
            if balance is not None:
                prop["mortgage_balance"] = float(balance)
            if emi is not None:
                prop["mortgage_emi"] = float(emi)
            if annual_rate is not None:
                prop["mortgage_interest_annual"] = float(annual_rate)
        out.append(prop)
    return out


def extract_goal_properties(ws: Worksheet) -> list[dict]:
    """Cols B-F (2-6), rows 64-81. Always treat property_amount as target_pv (per spec §7.5)."""
    out: list[dict] = []
    for col in range(2, 7):
        amount = ws.cell(row=65, column=col).value
        if amount is None or amount == "" or amount == 0:
            continue
        name = ws.cell(row=64, column=col).value
        if name is None:
            continue
        is_dp_only = bool(ws.cell(row=66, column=col).value)
        upfront = ws.cell(row=67, column=col).value
        goal_dt = ws.cell(row=69, column=col).value
        infl = ws.cell(row=72, column=col).value
        tenure = ws.cell(row=78, column=col).value
        rate = ws.cell(row=81, column=col).value
        prop = {
            "name": str(name),
            "target_pv": float(amount),
            "is_downpayment_only": is_dp_only,
        }
        if isinstance(goal_dt, datetime):
            prop["goal_date"] = goal_dt.date().isoformat()
        elif isinstance(goal_dt, date):
            prop["goal_date"] = goal_dt.isoformat()
        if upfront is not None and upfront != "":
            prop["upfront_amount"] = float(upfront)
        if infl is not None:
            prop["inflation_annual"] = float(infl)
        if tenure is not None and tenure != "":
            prop["mortgage_tenure_years"] = int(tenure)
        if rate is not None and rate != "":
            prop["mortgage_interest_annual"] = float(rate)
        out.append(prop)
    return out


def extract_custom_goals(ws: Worksheet) -> list[dict]:
    """Goals table rows 93-112; skip retirement and goal_property_* rows."""
    out: list[dict] = []
    for r in range(93, 113):
        name = ws.cell(row=r, column=2).value  # col B
        if not name:
            continue
        name_s = str(name).strip()
        if name_s == "retirement" or name_s.startswith("goal_property_"):
            continue
        amount_pv = ws.cell(row=r, column=7).value  # col G
        if amount_pv is None or amount_pv == "":
            continue
        goal_type = ws.cell(row=r, column=11).value  # col K
        goal_dt = ws.cell(row=r, column=10).value  # col J (Date)
        goal_dt_iso = None
        if isinstance(goal_dt, datetime):
            goal_dt_iso = goal_dt.date().isoformat()
        elif isinstance(goal_dt, date):
            goal_dt_iso = goal_dt.isoformat()
        if goal_dt_iso is None:
            continue
        # Map Excel goal_type strings to GoalType enum values
        gt = str(goal_type).strip() if goal_type else "custom"
        # Handle truncations / variants
        if gt.startswith("child_abroad"):
            gt = "child_abroad_education"
        elif gt.startswith("child_local"):
            gt = "child_local_education"
        elif gt.startswith("child_marriage"):
            gt = "child_marriage"
        elif gt == "":
            gt = "custom"
        out.append({
            "name": name_s,
            "goal_type": gt,
            "amount_pv": float(amount_pv),
            "goal_date": goal_dt_iso,
        })
    return out


def extract_one_offs(ws: Worksheet, start_row: int, end_row: int) -> list[dict]:
    """Generic one-off extractor. Cols: B=description, E=amount, F=Date."""
    out: list[dict] = []
    for r in range(start_row, end_row + 1):
        desc = ws.cell(row=r, column=2).value
        if not desc:
            continue
        amount = ws.cell(row=r, column=5).value
        dt = ws.cell(row=r, column=6).value
        if amount is None or dt is None:
            continue
        if isinstance(dt, datetime):
            dt_iso = dt.date().isoformat()
        elif isinstance(dt, date):
            dt_iso = dt.isoformat()
        else:
            continue
        out.append({
            "description": str(desc).strip(),
            "amount": float(amount),
            "date": dt_iso,
        })
    return out


def build_input(ws: Worksheet) -> dict:
    return {
        "assumptions": extract_assumptions(ws),
        "profile": extract_profile(ws),
        "retirement": extract_retirement(ws),
        "current_properties": extract_current_properties(ws),
        "goal_properties": extract_goal_properties(ws),
        "custom_goals": extract_custom_goals(ws),
        "one_off_inflows": extract_one_offs(ws, 133, 142),
        "one_off_outflows": extract_one_offs(ws, 118, 127),
        "detail_level": "full",
    }


# ---------------------------------------------------------------------------
# Output extraction (cell name → expected value)
# ---------------------------------------------------------------------------

def _get(ws: Worksheet, cell: str):
    """Coerce to a JSON-serializable scalar."""
    v = ws[cell].value
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return float(v) if not isinstance(v, bool) else v
    return str(v)


def build_expected(ws: Worksheet) -> dict:
    """Flat dict of Excel-cell → expected value. Keys mirror cell_mapping.md §2."""
    expected: dict = {}

    # 2a. headline
    for cell in ["B26", "B86", "B87", "B88", "B89", "L113", "M113", "O113", "S105"]:
        expected[cell] = _get(ws, cell)
    # closing_nfa lives at last annual row's col S — by convention row 214 (FY 2051-03-31)
    expected["S214"] = _get(ws, "S214")

    # 2b. retirement
    for cell in ["B39", "B42", "B43", "B44", "B45", "B46"]:
        expected[cell] = _get(ws, cell)

    # 2c. per-goal — rows 93..102 (goals 1..10), cols B/G/H/J/L/M/N/O
    goals = []
    for r in range(93, 103):
        name = ws.cell(row=r, column=2).value
        if not name:
            continue
        goals.append({
            "row": r,
            "name": str(name).strip(),
            "goal_date": _get(ws, f"J{r}"),
            "amount_pv": _get(ws, f"G{r}"),
            "amount_fv": _get(ws, f"H{r}"),
            "fund_today_pv": _get(ws, f"O{r}"),
            "funded_amount": _get(ws, f"M{r}"),
            "shortfall_fv": _get(ws, f"L{r}"),
            "expected_roi": _get(ws, f"N{r}"),
        })
    expected["_goals"] = goals

    # 2d. per-one-off-outflow — rows 118..120, cols B/E/F + per-outflow underfunded total at AS290..
    outflows = []
    underfunded_cols = ["AS", "AT", "AU", "AV", "AW", "AX", "AY", "AZ", "BA", "BB", "BC"]
    # AS290 = first slot (one_off_out_aggregate?). Looking at row 146/189:
    #   AS = amount_one_off_out (aggregate), AT..BC = per-goal slots
    # So per-one-off-outflow at row level lives in U column (totals U290 = 10M).
    # Per-goal underfunded at AT290..BC290.
    # Per-outflow underfunded is NOT broken out in row 290 — Q290 is the aggregate goal_cash_outflows.
    # We'll just record the descriptor + amount + Q290 aggregate.
    for r in range(118, 128):
        desc = ws.cell(row=r, column=2).value
        if not desc:
            continue
        outflows.append({
            "row": r,
            "description": str(desc).strip(),
            "date": _get(ws, f"F{r}"),
            "amount": _get(ws, f"E{r}"),
        })
    expected["_one_off_outflows"] = outflows
    expected["U290"] = _get(ws, "U290")  # total one-off outflows

    # 2e. fund_flow_summary — mini fund flow at S93..S99
    expected["S93"] = _get(ws, "S93")  # opening_nfa
    expected["S94"] = _get(ws, "S94")  # total_investments
    expected["S95"] = _get(ws, "S95")  # total_roi
    expected["S96"] = _get(ws, "S96")  # total_one_off_in
    expected["S97"] = _get(ws, "S97")  # -total_one_off_out (Excel sign)
    expected["S98"] = _get(ws, "S98")  # -total_goals_paid (Excel sign)
    expected["S99"] = _get(ws, "S99")  # -closing_nfa (Excel sign)

    # 2f. goal_property_details — cols B-F, rows 64-83
    props = []
    for col in range(2, 7):
        amt = ws.cell(row=65, column=col).value
        if amt is None or amt == "" or amt == 0:
            continue
        L = ws.cell(row=64, column=col).column_letter
        props.append({
            "col": L,
            "name": _get(ws, f"{L}64"),
            "target_pv": _get(ws, f"{L}65"),
            "is_downpayment_only": _get(ws, f"{L}66"),
            "upfront_amount": _get(ws, f"{L}67"),
            "goal_date": _get(ws, f"{L}69"),
            "target_fv": _get(ws, f"{L}75"),
            "payout_amount_fv": _get(ws, f"{L}76"),
            "mortgage_amount": _get(ws, f"{L}77"),
            "mortgage_tenure_years": _get(ws, f"{L}78"),
            "mortgage_payoff_date": _get(ws, f"{L}80"),
            "mortgage_interest_annual": _get(ws, f"{L}81"),
            "mortgage_emi_monthly": _get(ws, f"{L}83"),
        })
    expected["_goal_property_details"] = props

    # 2g. annual_cashflow — rows 190..214 (sl_no 1..25, 25 FYs total)
    annual = []
    for r in range(190, 215):
        if ws.cell(row=r, column=3).value is None:
            break
        annual.append({
            "row": r,
            "fy_end_date": _get(ws, f"C{r}"),
            "income": _get(ws, f"D{r}"),
            "income_tax": _get(ws, f"E{r}"),
            "household_expense": _get(ws, f"F{r}"),
            "savings_1": _get(ws, f"G{r}"),
            "existing_mortgage_emi_total": _get(ws, f"H{r}"),
            "goal_mortgage_emi_total": _get(ws, f"I{r}"),
            "savings_2": _get(ws, f"J{r}"),
            "investment_amount": _get(ws, f"M{r}"),
            "one_off_in": _get(ws, f"P{r}"),
            "one_off_out": _get(ws, f"U{r}"),
            "close_nfa": _get(ws, f"S{r}"),
        })
    expected["_annual_cashflow"] = annual

    # 2h. monthly spot-check rows
    monthly = []
    spot_rows = [147, 149, 161, 182]  # FY1-M1, FY1-M3, FY2-M12, last monthly
    for r in spot_rows:
        if ws.cell(row=r, column=2).value is None:
            continue
        monthly.append({
            "row": r,
            "month_end": _get(ws, f"B{r}"),
            "income": _get(ws, f"D{r}"),
            "income_tax": _get(ws, f"E{r}"),
            "household_expense": _get(ws, f"F{r}"),
            "savings_1": _get(ws, f"G{r}"),
            "existing_mortgage_emi_total": _get(ws, f"H{r}"),
            "goal_mortgage_emi_total": _get(ws, f"I{r}"),
            "savings_2": _get(ws, f"J{r}"),
            "savings_2_avg": _get(ws, f"K{r}"),
            # NFA tape
            "nfa_open": _get(ws, f"N{r}"),
            "regular_invest": _get(ws, f"M{r}"),
            "roi": _get(ws, f"O{r}"),
            "one_off_in": _get(ws, f"P{r}"),
            "goal_outflow_total": _get(ws, f"Q{r}"),
            "nfa_close": _get(ws, f"S{r}"),
        })
    expected["_monthly_spot"] = monthly

    return expected


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> int:
    if not EXCEL_PATH.exists():
        print(f"ERROR: Excel not found at {EXCEL_PATH}", file=sys.stderr)
        return 1
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    print(f"Loading {EXCEL_PATH}...")
    wb = load_workbook(EXCEL_PATH, data_only=True)
    if "Goal planning" not in wb.sheetnames:
        print(f"ERROR: 'Goal planning' sheet not found. Sheets: {wb.sheetnames}", file=sys.stderr)
        return 1
    ws = wb["Goal planning"]

    print("Extracting input...")
    inp = build_input(ws)
    inp_path = OUT_DIR / "input.json"
    inp_path.write_text(json.dumps(inp, indent=2, default=str))
    print(f"  → {inp_path}")
    print(f"    profile.annual_income={inp['profile']['annual_income']}")
    print(f"    profile.financial_assets={inp['profile']['financial_assets']}")
    print(f"    {len(inp['current_properties'])} current_properties, "
          f"{len(inp['goal_properties'])} goal_properties, "
          f"{len(inp['custom_goals'])} custom_goals")
    print(f"    {len(inp['one_off_inflows'])} one_off_in, "
          f"{len(inp['one_off_outflows'])} one_off_out")

    print("Extracting expected outputs...")
    exp = build_expected(ws)
    exp_path = OUT_DIR / "expected.json"
    exp_path.write_text(json.dumps(exp, indent=2, default=str))
    print(f"  → {exp_path}")
    print(f"    headline.B26 (NFA today)         = {exp['B26']}")
    print(f"    headline.M113 (total funded)     = {exp['M113']}")
    print(f"    headline.L113 (total shortfall)  = {exp['L113']}")
    print(f"    headline.S214 (closing_nfa)      = {exp['S214']}")
    print(f"    {len(exp['_goals'])} goals, {len(exp['_annual_cashflow'])} annual rows")

    return 0


if __name__ == "__main__":
    sys.exit(main())
